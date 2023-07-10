


import openpyxl        # library

class ReadData():

    @staticmethod
    def getRowCount(file,sheetName):                     # for complete the iteration if rowcount=3 then 3 iterations
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetName]
        return(sheet.max_row)               # row count

    @staticmethod
    def getColumnCount(file,sheetname):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetname]
        return(sheet.max_column)           # column count

    @staticmethod
    def readData(file,sheetname,rownum,columnnum):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetname]
        return sheet.cell(row=rownum,column=columnnum).value

    @staticmethod
    def writeData(file,sheetname,rownum,columnnum):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetname]
        return sheet.cell(row=rownum,column=columnnum).value
        workbook.save(file)



#------------------------------------------------------------------------------------------












